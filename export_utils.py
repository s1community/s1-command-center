"""
Export utilities — generates beautiful HTML and Excel reports from table data.
"""
import json
import os
import re
from collections import Counter
from datetime import datetime
from tkinter import filedialog, messagebox
from typing import Optional

from app import cli_log


# ═══════════════════════════════════════════════════════════════════════
#  HTML Report
# ═══════════════════════════════════════════════════════════════════════

_CSS = """
* { margin: 0; padding: 0; box-sizing: border-box; }
body {
    font-family: 'Segoe UI', -apple-system, BlinkMacSystemFont, sans-serif;
    background: #0d0d1a; color: #e0e0e0; padding: 40px;
}
.header {
    background: linear-gradient(135deg, #1a1a2e 0%, #16213e 100%);
    border-radius: 16px; padding: 32px 40px; margin-bottom: 32px;
    border: 1px solid #2d2d44;
}
.header h1 { font-size: 28px; color: #fff; margin-bottom: 4px; }
.header .subtitle { color: #888; font-size: 14px; }
.header .meta { color: #666; font-size: 12px; margin-top: 12px; }
.stats {
    display: flex; gap: 16px; margin-bottom: 28px; flex-wrap: wrap;
}
.stat-card {
    background: #1a1a2e; border: 1px solid #2d2d44; border-radius: 12px;
    padding: 20px 28px; min-width: 160px;
}
.stat-card .label { font-size: 12px; color: #888; text-transform: uppercase;
    letter-spacing: 1px; margin-bottom: 4px; }
.stat-card .value { font-size: 28px; font-weight: 700; color: #00b894; }
.stat-card .value.accent { color: #e94560; }
.stat-card .value.warn { color: #fdcb6e; }
table {
    width: 100%; border-collapse: collapse; background: #1a1a2e;
    border-radius: 12px; overflow: hidden; border: 1px solid #2d2d44;
}
thead th {
    background: #16213e; color: #aaa; font-size: 11px;
    text-transform: uppercase; letter-spacing: 1px; padding: 14px 16px;
    text-align: left; border-bottom: 2px solid #2d2d44;
    position: sticky; top: 0;
}
tbody td {
    padding: 10px 16px; border-bottom: 1px solid #222238;
    font-size: 13px; max-width: 300px; overflow: hidden;
    text-overflow: ellipsis; white-space: nowrap;
}
tbody tr:hover { background: #222238; }
tbody tr:nth-child(even) { background: #151528; }
.badge {
    display: inline-block; padding: 2px 10px; border-radius: 12px;
    font-size: 11px; font-weight: 600;
}
.badge-green { background: #00b89422; color: #00b894; }
.badge-red { background: #e9456022; color: #e94560; }
.badge-yellow { background: #fdcb6e22; color: #fdcb6e; }
.badge-blue { background: #0984e322; color: #74b9ff; }
.footer {
    text-align: center; color: #444; font-size: 11px;
    margin-top: 32px; padding-top: 16px; border-top: 1px solid #222;
}
"""

# Fields that get badge styling
_BADGE_MAP = {
    "true": "badge-green", "false": "badge-red",
    "active": "badge-green", "infected": "badge-red",
    "mitigated": "badge-green", "not_mitigated": "badge-red",
    "suspicious": "badge-yellow", "malicious": "badge-red",
    "critical": "badge-red", "high": "badge-red",
    "medium": "badge-yellow", "low": "badge-blue",
    "finished": "badge-green", "running": "badge-yellow",
    "enabled": "badge-green", "disabled": "badge-red",
}


def _badge(val: str) -> str:
    v = str(val).lower().strip()
    cls = _BADGE_MAP.get(v, "")
    if cls:
        return f'<span class="badge {cls}">{val}</span>'
    return str(val)


def _cell(val) -> str:
    if isinstance(val, (dict, list)):
        return json.dumps(val, default=str)[:120]
    s = str(val)
    return s[:200] if len(s) > 200 else s


def generate_html(title: str, columns: list[str], rows: list[dict],
                  stats: Optional[list[dict]] = None,
                  subtitle: str = "") -> str:
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    stats_html = ""
    if stats:
        cards = ""
        for s in stats:
            cls = s.get("class", "")
            cards += (f'<div class="stat-card"><div class="label">{s["label"]}</div>'
                      f'<div class="value {cls}">{s["value"]}</div></div>')
        stats_html = f'<div class="stats">{cards}</div>'

    # Table
    th = "".join(f"<th>{c}</th>" for c in columns)
    tr_list = []
    for row in rows:
        tds = "".join(f"<td>{_badge(_cell(row.get(c, '')))}</td>" for c in columns)
        tr_list.append(f"<tr>{tds}</tr>")
    tbody = "\n".join(tr_list)

    return f"""<!DOCTYPE html>
<html lang="en"><head><meta charset="utf-8">
<title>{title} — S1 Command Center Report</title>
<style>{_CSS}</style></head><body>
<div class="header">
  <h1>{title}</h1>
  <div class="subtitle">{subtitle or 'S1 Command Center Report'}</div>
  <div class="meta">Generated {now} &bull; {len(rows)} records</div>
</div>
{stats_html}
<table><thead><tr>{th}</tr></thead><tbody>{tbody}</tbody></table>
<div class="footer">S1 Command Center &bull; Made by Ran Jacobi &bull; Generated {now}</div>
</body></html>"""


# ═══════════════════════════════════════════════════════════════════════
#  Backup redaction (safe-to-share copies)
# ═══════════════════════════════════════════════════════════════════════

REDACTED = "***REDACTED***"

# A key is treated as secret if its name contains any of these (case-insensitive).
# Backups embed real secrets in the settings blocks — SMTP relay passwords, AD
# bind credentials, syslog tokens, SSO client secrets / private keys, webhook
# auth headers — so a backup JSON shared on a ticket or chat is a credential
# leak. Redaction produces a sanitised COPY for sharing; the working backup
# used for restore is never modified.
_SECRET_KEY_PARTS = (
    "password", "passwd", "passphrase", "secret", "token", "apikey",
    "api_key", "privatekey", "private_key", "clientsecret", "client_secret",
    "bindpassword", "bind_password", "credential", "authorization", "bearer",
)


def _is_secret_key(key) -> bool:
    k = str(key).lower().replace("-", "").replace("_", "")
    return any(p.replace("_", "") in k for p in _SECRET_KEY_PARTS)


def _redact_obj(obj, counter):
    """Recursively return a redacted deep copy, counting masked values."""
    if isinstance(obj, dict):
        out = {}
        for k, v in obj.items():
            if _is_secret_key(k) and v not in (None, "", [], {}):
                out[k] = REDACTED
                counter[0] += 1
            else:
                out[k] = _redact_obj(v, counter)
        return out
    if isinstance(obj, list):
        return [_redact_obj(i, counter) for i in obj]
    return obj


def count_backup_secrets(nodes) -> int:
    """How many secret values a backup contains (0 = safe to share as-is)."""
    counter = [0]
    _redact_obj(nodes or [], counter)
    return counter[0]


def redact_backup(nodes):
    """Return (redacted_copy, count) — a deep copy of the backup with every
    secret-looking value masked. Input is not mutated."""
    counter = [0]
    redacted = _redact_obj(nodes or [], counter)
    return redacted, counter[0]


# ═══════════════════════════════════════════════════════════════════════
#  Migration Manifest (for the PSO ticket-closing workflow)
# ═══════════════════════════════════════════════════════════════════════

def build_migration_manifest(meta: dict, results: list[dict]) -> dict:
    """Turn a Migration Validation result set (meta + results, the structure
    produced by ValidationPage._run_validation) into a structured, portable
    manifest of what was migrated and how it verified.

    Pure function — no GUI/network. Safe to unit-test."""
    meta = meta or {}
    results = results or []

    n = len(results)
    missing = sum(1 for r in results if not r.get("matched"))
    with_diffs = sum(1 for r in results
                     if r.get("matched") and r.get("diffs", 0) > 0)
    identical = n - missing - with_diffs
    total_diffs = sum(r.get("diffs", 0) for r in results if r.get("matched"))

    if n == 0:
        status = "incomplete"
    elif missing == 0 and with_diffs == 0:
        status = "verified"
    else:
        status = "differences"

    nodes = []
    for r in results:
        if not r.get("matched"):
            outcome = "missing"
            diffs = []
        elif r.get("diffs", 0) == 0:
            outcome = "identical"
            diffs = []
        else:
            outcome = "differences"
            diffs = [
                {
                    "element": x.get("cat"),
                    "src": x.get("src"),
                    "dst": x.get("dst"),
                    "missing": x.get("missing", []),
                    "extra": x.get("extra", []),
                }
                for x in r.get("rows", []) if x.get("status") == "diff"
            ]
        nodes.append({
            "type": r.get("type"),
            "path": r.get("path"),
            "result": outcome,
            "differences": diffs,
        })

    return {
        "tool": "S1 Command Center",
        "kind": "migration-manifest",
        "version": 1,
        "generatedAt": meta.get("when"),
        "source": meta.get("src_url"),
        "destination": meta.get("dst_url"),
        "levels": meta.get("levels", []),
        "scope": {
            "source": meta.get("src_filters", {}),
            "destination": meta.get("dst_filters", {}),
        },
        "summary": {
            "nodesCompared": n,
            "identical": identical,
            "withDifferences": with_diffs,
            "missingOnDestination": missing,
            "totalDifferences": total_diffs,
            "status": status,
        },
        "nodes": nodes,
    }


def manifest_to_pso_comment(manifest: dict) -> str:
    """Render a migration manifest as a Markdown comment ready for the PSO
    Jira ticket-closing workflow ('done with PSO-XXX'). Mirrors the agreed
    Migration Summary template so the comment can be posted as-is."""
    m = manifest or {}
    s = m.get("summary", {})
    when = (m.get("generatedAt") or "")[:10] or "—"
    status = s.get("status")

    if status == "verified":
        headline = "Migration completed successfully ✅"
        status_line = "All settings transferred and verified."
    elif status == "differences":
        headline = "Migration completed — review needed ⚠️"
        status_line = (
            f"{s.get('totalDifferences', 0)} difference(s) across "
            f"{s.get('withDifferences', 0)} node(s); "
            f"{s.get('missingOnDestination', 0)} scope(s) missing on "
            f"destination. See details below.")
    else:
        headline = "Migration validation incomplete"
        status_line = "No nodes were compared — re-run validation."

    src_scope = (m.get("scope", {}).get("source") or {})
    scope_txt = (f"account: {src_scope.get('account') or 'all'} · "
                 f"site: {src_scope.get('site') or 'all'}")
    levels = ", ".join(m.get("levels", [])) or "—"

    lines = [
        headline,
        "",
        "**Migration Summary**",
        f"- **Source:** {m.get('source') or '?'}",
        f"- **Destination:** {m.get('destination') or '?'}",
        f"- **Completed:** {when}",
        f"- **Scope:** levels [{levels}] · {scope_txt}",
        f"- **Validation:** {s.get('nodesCompared', 0)} node(s) compared — "
        f"{s.get('identical', 0)} identical, "
        f"{s.get('withDifferences', 0)} with differences, "
        f"{s.get('missingOnDestination', 0)} missing on destination",
        f"- **Status:** {status_line}",
    ]

    # Per-node difference detail (only when there is something to flag).
    flagged = [nd for nd in m.get("nodes", [])
               if nd.get("result") in ("differences", "missing")]
    if flagged:
        lines += ["", "**Items needing attention**"]
        for nd in flagged:
            if nd.get("result") == "missing":
                lines.append(
                    f"- `{nd.get('path')}` — scope missing on destination "
                    f"(renamed or not created)")
                continue
            for d in nd.get("differences", []):
                detail = []
                if d.get("missing"):
                    miss = d["missing"]
                    shown = ", ".join(miss[:8])
                    if len(miss) > 8:
                        shown += f" (+{len(miss) - 8} more)"
                    detail.append(f"missing: {shown}")
                if d.get("extra"):
                    detail.append(f"{len(d['extra'])} extra on dest")
                tail = f" ({'; '.join(detail)})" if detail else ""
                lines.append(
                    f"- `{nd.get('path')}` / {d.get('element')}: "
                    f"{d.get('src')} → {d.get('dst')}{tail}")

    return "\n".join(lines)


# ═══════════════════════════════════════════════════════════════════════
#  Excel Report (via openpyxl)
# ═══════════════════════════════════════════════════════════════════════

def generate_excel(path: str, title: str, columns: list[str],
                   rows: list[dict]):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]

    # Colors
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e",
                              fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Segoe UI", size=10, color="333333")
    alt_fill = PatternFill(start_color="F5F6FA", end_color="F5F6FA",
                           fill_type="solid")
    border = Border(
        bottom=Side(style="thin", color="E0E0E0"))

    # Title row
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=max(len(columns), 1))
    cell = ws.cell(row=1, column=1, value=f"{title} — S1 Command Center Report")
    cell.font = Font(name="Segoe UI", size=14, bold=True, color="1a1a2e")
    cell.alignment = Alignment(horizontal="left")

    ws.merge_cells(start_row=2, start_column=1,
                   end_row=2, end_column=max(len(columns), 1))
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    cell2 = ws.cell(row=2, column=1,
                    value=f"Generated {now}  •  {len(rows)} records")
    cell2.font = Font(name="Segoe UI", size=9, color="888888")

    # Header row
    for j, col in enumerate(columns, 1):
        c = ws.cell(row=4, column=j, value=col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="left")

    # Data rows
    for i, row in enumerate(rows):
        r = i + 5
        fill = alt_fill if i % 2 == 0 else None
        for j, col in enumerate(columns, 1):
            val = row.get(col, "")
            if isinstance(val, (dict, list)):
                val = json.dumps(val, default=str)
            c = ws.cell(row=r, column=j, value=val)
            c.font = data_font
            c.border = border
            if fill:
                c.fill = fill

    # Auto-width
    for j, col in enumerate(columns, 1):
        max_len = len(col)
        for i, row in enumerate(rows[:200]):
            val = str(row.get(col, ""))
            max_len = max(max_len, min(len(val), 50))
        ws.column_dimensions[ws.cell(row=4, column=j).column_letter].width = \
            max_len + 4

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{ws.cell(row=4, column=len(columns)).column_letter}4"
    wb.save(path)


# ═══════════════════════════════════════════════════════════════════════
#  STAR (custom detection) rules — detailed Excel workbook
# ═══════════════════════════════════════════════════════════════════════

_XL_FONT = "Segoe UI"
# Excel rejects most C0 control characters and caps a cell at 32,767 chars.
_XL_ILLEGAL_RE = re.compile(r"[\000-\010\013\014\016-\037]")
_XL_MAX_CHARS = 32000

# (header, rule key, column width, kind)
STAR_COLUMNS = [
    ("Rule Name",          "name",              38, "text"),
    ("Description",        "description",       46, "wrap"),
    ("Status",             "status",            12, "status"),
    ("Severity",           "severity",          11, "severity"),
    ("Scope",              "scope",             10, "scope"),
    ("Scope Path",         "scopeName",         34, "text"),
    ("Account",            "accountName",       22, "text"),
    ("Site",               "siteName",          22, "text"),
    ("Query Type",         "queryType",         13, "text"),
    ("Query Language",     "queryLang",         14, "text"),
    ("Detection Query",    "s1ql",              70, "wrap"),
    ("Treat as Threat",    "treatAsThreat",     15, "text"),
    ("Network Quarantine", "networkQuarantine", 18, "bool"),
    ("Active Response",    "activeResponse",    24, "wrap"),
    ("Expiration Mode",    "expirationMode",    15, "text"),
    ("Expires",            "expiration",        17, "date"),
    ("Expired",            "expired",           9,  "bool"),
    ("Alerts Generated",   "generatedAlerts",   15, "num"),
    ("Last Alert",         "lastAlertTime",     17, "date"),
    ("Created",            "createdAt",         17, "date"),
    ("Created By",         "creator",           20, "text"),
    ("Last Updated",       "updatedAt",         17, "date"),
    ("Updated By",         "updater",           20, "text"),
    ("Rule ID",            "id",                22, "text"),
]

# value -> (fill, font colour)
_STAR_SEVERITY_STYLE = {
    "critical":      ("FFE0E0", "B02020"),
    "high":          ("FFEBDC", "C2560A"),
    "medium":        ("FFF6DC", "9C6F00"),
    "low":           ("E6F0FB", "1B5FA8"),
    "suspicious":    ("FFF6DC", "9C6F00"),
    "info":          ("EFEFF4", "55556A"),
    "informational": ("EFEFF4", "55556A"),
}
_STAR_STATUS_STYLE = {
    "active":   ("E3F7EC", "1B7F4F"),
    "draft":    ("EDEDF2", "555566"),
    "disabled": ("FDEAE7", "B02020"),
}
_STAR_SCOPE_STYLE = {
    "global":  ("EDE6F8", "5B34B0"),
    "tenant":  ("EDE6F8", "5B34B0"),
    "account": ("E2F0FD", "13599C"),
    "site":    ("E6F5E9", "27702F"),
    "group":   ("FFF0DC", "A85B00"),
}

_STAR_SCOPE_ORDER = {"global": 0, "tenant": 0, "account": 1, "site": 2,
                     "group": 3}


def _xl_safe(value):
    """Strip characters Excel refuses and cap the cell length."""
    if not isinstance(value, str):
        return value
    value = _XL_ILLEGAL_RE.sub("", value)
    if len(value) > _XL_MAX_CHARS:
        value = value[:_XL_MAX_CHARS] + "… (truncated)"
    return value


def _fmt_dt(val) -> str:
    """'2024-01-31T09:15:00.000000Z' -> '2024-01-31 09:15'."""
    if not val:
        return ""
    s = str(val)
    try:
        return datetime.fromisoformat(
            s.replace("Z", "+00:00")).strftime("%Y-%m-%d %H:%M")
    except Exception:
        return s[:19].replace("T", " ")


def _fmt_active_response(val) -> str:
    """Render activeResponse — a plain on/off flag on some consoles, an object
    of individual actions on others."""
    if val is None or val == "":
        return ""
    if isinstance(val, bool):
        return "Yes" if val else "No"
    if isinstance(val, dict):
        on = [k for k, v in val.items() if v is True]
        if on:
            return ", ".join(on)
        kept = {k: v for k, v in val.items()
                if v not in (None, False, "", [], {})}
        return json.dumps(kept, default=str) if kept else "No"
    return str(val)


def _star_cell(rule: dict, key: str, kind: str):
    val = rule.get(key)
    if key == "activeResponse":
        return _fmt_active_response(val)
    if kind == "date":
        return _fmt_dt(val)
    if kind == "bool":
        if val is None:
            return ""
        return "Yes" if val else "No"
    if kind == "num":
        return val if isinstance(val, (int, float)) else (val or 0)
    if isinstance(val, (dict, list)):
        return _xl_safe(json.dumps(val, default=str))
    if val is None:
        return ""
    return _xl_safe(str(val))


def _star_sort_key(r: dict):
    return (
        str(r.get("accountName") or "").lower(),
        _STAR_SCOPE_ORDER.get(str(r.get("scope") or "").lower(), 9),
        str(r.get("siteName") or "").lower(),
        str(r.get("name") or "").lower(),
    )


def count_star_scope_duplicates(rules) -> int:
    """How many SITE-scoped rules share a name with an account rule in the
    same account. Surfaced on the summary sheet because that is the signature
    of the pre-2.1.9 restore bug that copied account rules to every site."""
    acct = {(str(r.get("accountId") or ""), r.get("name"))
            for r in (rules or [])
            if str(r.get("scope") or "").lower() == "account"}
    return sum(1 for r in (rules or [])
               if str(r.get("scope") or "").lower() == "site"
               and (str(r.get("accountId") or ""), r.get("name")) in acct)


def generate_star_rules_excel(path: str, rules: list,
                              meta: Optional[dict] = None) -> int:
    """Write a polished, filterable workbook of STAR custom detection rules.

    'Summary'    — headline totals plus breakdowns by scope, status, severity
                   and account.
    'STAR Rules' — every rule with all customer-relevant fields, frozen
                   header, auto-filter and colour-coded scope/status/severity
                   so it is readable without any further formatting work.

    Returns the number of rules written."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    meta = meta or {}
    rules = sorted(rules or [], key=_star_sort_key)
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    ink = "1A1A2E"
    row_border = Border(bottom=Side(style="thin", color="E3E3EC"))
    head_border = Border(bottom=Side(style="medium", color=ink))

    wb = Workbook()

    # ── Sheet 1: Summary ────────────────────────────────────────────────
    s = wb.active
    s.title = "Summary"
    s.sheet_view.showGridLines = False
    for col, width in (("A", 3), ("B", 38), ("C", 14), ("D", 12)):
        s.column_dimensions[col].width = width

    t = s.cell(row=2, column=2, value="STAR Custom Detection Rules")
    t.font = Font(name=_XL_FONT, size=18, bold=True, color=ink)
    sub = s.cell(row=3, column=2, value="S1 Command Center export")
    sub.font = Font(name=_XL_FONT, size=10, color="8A8A9A")

    r = 5
    for label, value in (
        ("Console", meta.get("console") or "—"),
        ("Account filter", meta.get("account_filter") or "(all accounts)"),
        ("Site filter", meta.get("site_filter") or "(all sites)"),
        ("Generated", now),
        ("Total rules", len(rules)),
    ):
        lc = s.cell(row=r, column=2, value=label)
        lc.font = Font(name=_XL_FONT, size=10, color="8A8A9A")
        vc = s.cell(row=r, column=3, value=value)
        vc.font = Font(name=_XL_FONT, size=10, bold=True, color=ink)
        r += 1

    dupes = count_star_scope_duplicates(rules)
    dl = s.cell(row=r, column=2,
                value="Site rules duplicating an account rule")
    dl.font = Font(name=_XL_FONT, size=10, color="8A8A9A")
    dv = s.cell(row=r, column=3, value=dupes)
    dv.font = Font(name=_XL_FONT, size=10, bold=True,
                   color="B02020" if dupes else "1B7F4F")
    r += 2

    def _breakdown(row, heading, counter, style_map=None):
        for col, text in ((2, heading), (3, "Count")):
            h = s.cell(row=row, column=col, value=text)
            h.font = Font(name=_XL_FONT, size=11, bold=True, color=ink)
            h.border = head_border
        row += 1
        total = sum(counter.values()) or 1
        for label, count in counter.most_common():
            lc = s.cell(row=row, column=2, value=str(label or "—"))
            lc.font = Font(name=_XL_FONT, size=10, color="333344")
            lc.border = row_border
            cc = s.cell(row=row, column=3, value=count)
            cc.font = Font(name=_XL_FONT, size=10, bold=True, color="333344")
            cc.border = row_border
            pc = s.cell(row=row, column=4, value=count / total)
            pc.number_format = "0.0%"
            pc.font = Font(name=_XL_FONT, size=10, color="8A8A9A")
            pc.border = row_border
            sty = style_map.get(str(label or "").lower()) if style_map else None
            if sty:
                lc.fill = PatternFill("solid", start_color=sty[0],
                                      end_color=sty[0])
                lc.font = Font(name=_XL_FONT, size=10, bold=True,
                               color=sty[1])
            row += 1
        return row + 1

    r = _breakdown(r, "By scope",
                   Counter(str(x.get("scope") or "—").lower() for x in rules),
                   _STAR_SCOPE_STYLE)
    r = _breakdown(r, "By status",
                   Counter(str(x.get("status") or "—") for x in rules),
                   _STAR_STATUS_STYLE)
    r = _breakdown(r, "By severity",
                   Counter(str(x.get("severity") or "—") for x in rules),
                   _STAR_SEVERITY_STYLE)
    r = _breakdown(r, "By account",
                   Counter(str(x.get("accountName") or "—") for x in rules))

    # ── Sheet 2: the rules ──────────────────────────────────────────────
    ws = wb.create_sheet("STAR Rules")
    ws.sheet_view.showGridLines = False
    last_col = get_column_letter(len(STAR_COLUMNS))

    ws.merge_cells(f"A1:{last_col}1")
    tc = ws.cell(row=1, column=1, value="STAR Custom Detection Rules")
    tc.font = Font(name=_XL_FONT, size=15, bold=True, color=ink)
    tc.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26

    ws.merge_cells(f"A2:{last_col}2")
    sc = ws.cell(row=2, column=1,
                 value=f"{meta.get('console') or 'console'}  •  "
                       f"{len(rules)} rule(s)  •  generated {now}")
    sc.font = Font(name=_XL_FONT, size=9, color="8A8A9A")

    header_fill = PatternFill("solid", start_color=ink, end_color=ink)
    for j, (title, _key, width, _kind) in enumerate(STAR_COLUMNS, 1):
        c = ws.cell(row=4, column=j, value=title)
        c.font = Font(name=_XL_FONT, size=10, bold=True, color="FFFFFF")
        c.fill = header_fill
        c.alignment = Alignment(horizontal="left", vertical="center",
                                wrap_text=True)
        ws.column_dimensions[get_column_letter(j)].width = width
    ws.row_dimensions[4].height = 26

    alt = PatternFill("solid", start_color="F7F8FC", end_color="F7F8FC")
    for i, rule in enumerate(rules):
        row = i + 5
        banded = (i % 2 == 1)
        for j, (_title, key, _w, kind) in enumerate(STAR_COLUMNS, 1):
            c = ws.cell(row=row, column=j, value=_star_cell(rule, key, kind))
            c.font = Font(name=_XL_FONT, size=10, color="2C2C3A")
            c.border = row_border
            c.alignment = Alignment(horizontal="left", vertical="top",
                                    wrap_text=(kind == "wrap"))
            if banded:
                c.fill = alt
            if kind == "num":
                c.alignment = Alignment(horizontal="right", vertical="top")
            sty = None
            if kind == "severity":
                sty = _STAR_SEVERITY_STYLE.get(
                    str(rule.get("severity") or "").lower())
            elif kind == "status":
                sty = _STAR_STATUS_STYLE.get(
                    str(rule.get("status") or "").lower())
            elif kind == "scope":
                sty = _STAR_SCOPE_STYLE.get(
                    str(rule.get("scope") or "").lower())
            if sty:
                c.fill = PatternFill("solid", start_color=sty[0],
                                     end_color=sty[0])
                c.font = Font(name=_XL_FONT, size=10, bold=True, color=sty[1])
                c.alignment = Alignment(horizontal="center", vertical="top")

    ws.freeze_panes = "B5"
    ws.auto_filter.ref = f"A4:{last_col}{4 + len(rules)}"

    wb.save(path)
    return len(rules)


# ═══════════════════════════════════════════════════════════════════════
#  Unified export dialog
# ═══════════════════════════════════════════════════════════════════════

def export_report(title: str, columns: list[str], rows: list[dict],
                  stats: Optional[list[dict]] = None,
                  subtitle: str = ""):
    """Show save dialog and export as HTML or Excel based on user choice."""
    if not rows:
        messagebox.showwarning("No Data", "Nothing to export — load data first.")
        return

    ts = datetime.now().strftime("%Y%m%d-%H%M")
    safe_title = title.lower().replace(" ", "-").replace("&", "and")
    path = filedialog.asksaveasfilename(
        title=f"Export {title}",
        initialfile=f"s1-{safe_title}-{ts}",
        defaultextension=".html",
        filetypes=[
            ("HTML Report", "*.html"),
            ("Excel Workbook", "*.xlsx"),
            ("JSON Data", "*.json"),
        ],
    )
    if not path:
        return

    try:
        ext = os.path.splitext(path)[1].lower()
        if ext == ".xlsx":
            generate_excel(path, title, columns, rows)
        elif ext == ".json":
            with open(path, "w") as f:
                json.dump(rows, f, indent=2, default=str)
        else:
            html = generate_html(title, columns, rows, stats=stats,
                                 subtitle=subtitle)
            with open(path, "w") as f:
                f.write(html)

        cli_log(f"Exported {len(rows)} records → {os.path.basename(path)}",
                "success")
        cli_log(f"File saved to: {path}", "info")
    except Exception as e:
        cli_log(f"Export error: {e}", "error")
        messagebox.showerror("Export Error", str(e))
