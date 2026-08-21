"""Tests for the STAR custom detection rule Excel export.

The workbook is handed to customers, so the value formatting (dates, the
on/off vs object `activeResponse` shape, booleans) and the sheet structure
are pinned here.
"""
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from export_utils import (  # noqa: E402
    STAR_COLUMNS,
    _fmt_active_response,
    _fmt_dt,
    _star_cell,
    _star_sort_key,
    count_star_scope_duplicates,
    generate_star_rules_excel,
)


def _rule(**kw):
    base = {
        "id": "1", "name": "Rule", "description": "d", "scope": "account",
        "scopeName": "Acct", "accountId": "A1", "accountName": "Acct",
        "siteId": None, "siteName": None, "status": "Active",
        "severity": "High", "queryType": "events", "queryLang": "1.0",
        "s1ql": 'event.type = "Process Creation"', "treatAsThreat": "Suspicious",
        "networkQuarantine": False, "activeResponse": True,
        "expirationMode": "Permanent", "expiration": None, "expired": False,
        "generatedAlerts": 3, "lastAlertTime": None,
        "createdAt": "2026-01-31T09:15:00.000000Z", "creator": "a@b.com",
        "updatedAt": "2026-02-01T10:00:00.000000Z", "updater": "a@b.com",
    }
    base.update(kw)
    return base


# ── value formatting ────────────────────────────────────────────────────

def test_fmt_dt_normalises_iso_and_tolerates_junk():
    assert _fmt_dt("2026-01-31T09:15:00.000000Z") == "2026-01-31 09:15"
    assert _fmt_dt(None) == ""
    assert _fmt_dt("") == ""
    # unparseable input still degrades to something readable
    assert _fmt_dt("not-a-date") == "not-a-date"


def test_active_response_handles_bool_and_object_shapes():
    # some consoles return a plain flag...
    assert _fmt_active_response(True) == "Yes"
    assert _fmt_active_response(False) == "No"
    # ...others an object of individual actions
    assert _fmt_active_response(
        {"killProcess": True, "quarantine": True, "remediate": False}) == \
        "killProcess, quarantine"
    assert _fmt_active_response(None) == ""


def test_bool_columns_render_yes_no_not_true_false():
    assert _star_cell(_rule(networkQuarantine=True),
                      "networkQuarantine", "bool") == "Yes"
    assert _star_cell(_rule(networkQuarantine=False),
                      "networkQuarantine", "bool") == "No"
    assert _star_cell(_rule(networkQuarantine=None),
                      "networkQuarantine", "bool") == ""


def test_missing_values_never_render_the_string_none():
    for _hdr, key, _w, kind in STAR_COLUMNS:
        assert _star_cell({}, key, kind) in ("", 0)


def test_long_query_is_capped_to_excels_cell_limit():
    val = _star_cell(_rule(s1ql="x" * 40000), "s1ql", "wrap")
    assert len(val) < 32767


# ── sorting + duplicate metric ──────────────────────────────────────────

def test_sorted_account_rules_come_before_their_site_rules():
    rules = [
        _rule(id="s", scope="site", siteName="S1", name="b"),
        _rule(id="a", scope="account", name="a"),
    ]
    assert [r["id"] for r in sorted(rules, key=_star_sort_key)] == ["a", "s"]


def test_duplicate_metric_counts_site_copies_of_account_rules():
    rules = [
        _rule(id="1", scope="account", name="Dup"),
        _rule(id="2", scope="site", name="Dup", siteName="S1"),
        _rule(id="3", scope="site", name="Dup", siteName="S2"),
        _rule(id="4", scope="site", name="Unique", siteName="S1"),
    ]
    assert count_star_scope_duplicates(rules) == 2
    assert count_star_scope_duplicates([]) == 0


# ── workbook structure ──────────────────────────────────────────────────

def test_workbook_has_both_sheets_and_a_row_per_rule(tmp_path):
    from openpyxl import load_workbook

    rules = [_rule(id=str(i), name=f"R{i}") for i in range(5)]
    path = tmp_path / "star.xlsx"
    assert generate_star_rules_excel(str(path), rules,
                                     meta={"console": "https://c"}) == 5

    wb = load_workbook(path)
    assert wb.sheetnames == ["Summary", "STAR Rules"]

    ws = wb["STAR Rules"]
    assert [c.value for c in ws[4]] == [c[0] for c in STAR_COLUMNS]
    # header on row 4, one data row per rule
    assert ws.max_row == 4 + len(rules)
    # filters + frozen header are what make it usable for a customer
    assert ws.auto_filter.ref == f"A4:X{4 + len(rules)}"
    assert ws.freeze_panes == "B5"


def test_workbook_handles_zero_rules(tmp_path):
    path = tmp_path / "empty.xlsx"
    assert generate_star_rules_excel(str(path), [], meta={}) == 0
    assert path.exists()
